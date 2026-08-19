import csv
import re
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

Error_whitelist = [1020, 1064, 1342, 1376, 20287, 20291, 20292, 20317, 20318, 20319, 20320, 20321, 20322]
Have_Log_Server_list = ["WCDCDHCP01A","WCDCDHCP02A","WCDCDHCP03A","WKWCDHCP01","WKWCDHCP02","WKCCDHCP01","WKCCDHCP02","WNTWDHCP01","WNTWDHCP02","WKECDHCP01","WKECDHCP02","WHKWDHCP01","WHKWDHCP02","WHKEDHCP01","WHKEDHCP02","WHAHDHCP01","WHAHDHCP02","WNTEDHCP01","WNTEDHCP02"]


def _parse_csv_date(value):
	if not value:
		return None
	text = str(value).strip()
	# Normalize 1-digit month/day into zero-padded values (e.g. 8/3/2026 -> 08/03/2026).
	slash_date_match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})(.*)$", text)
	if slash_date_match:
		month, day, year, suffix = slash_date_match.groups()
		text = f"{int(month):02d}/{int(day):02d}/{year}{suffix}"
	date_formats = [
		"%Y-%m-%d %H:%M:%S",
		"%Y-%m-%d %H:%M",
		"%Y-%m-%d %I:%M:%S %p",
		"%Y-%m-%d %I:%M %p",
		"%Y/%m/%d %H:%M:%S",
		"%Y/%m/%d %H:%M",
		"%Y/%m/%d %I:%M:%S %p",
		"%Y/%m/%d %I:%M %p",
		"%m/%d/%Y %H:%M:%S",
		"%m/%d/%Y %H:%M",
		"%m/%d/%Y %I:%M:%S %p",
		"%m/%d/%Y %I:%M %p",
		"%d/%m/%Y %H:%M:%S",
		"%d/%m/%Y %H:%M",
		"%d/%m/%Y %I:%M:%S %p",
		"%d/%m/%Y %I:%M %p",
		"%Y-%m-%d",
		"%Y/%m/%d",
	]
	for fmt in date_formats:
		try:
			return datetime.strptime(text, fmt).date()
		except ValueError:
			continue

	# Fallback for other locale/system datetime strings.
	try:
		return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
	except ValueError:
		pass

	try:
		return datetime.strptime(text, "%c").date()
	except ValueError:
		pass
	return None


def _get_event_id(row):
	raw = row.get("Event ID") or row.get("EventID") or row.get("Event Id")
	if raw is None:
		return None
	try:
		return int(str(raw).strip())
	except ValueError:
		return None


def _get_row_date(row):
	date_text = row.get("Date and TIme") or row.get("Date and Time") or row.get("Date")
	return _parse_csv_date(date_text)


def _merge_task_category_overflow(row):
	# DictReader stores extra comma-split values in row[None]. Merge them back.
	extra_values = row.get(None)
	if not extra_values:
		return row

	if isinstance(extra_values, list):
		extra_text = ",".join(str(v) for v in extra_values if v is not None).strip()
	else:
		extra_text = str(extra_values).strip()

	task_key = "Task Category"
	for key in row.keys():
		if isinstance(key, str) and key.strip().lower() == "task category":
			task_key = key
			break

	current_value = str(row.get(task_key) or "").strip()
	if current_value.lower() == "none":
		current_value = ""
	if extra_text:
		row[task_key] = f"{current_value},{extra_text}" if current_value else extra_text

	row.pop(None, None)
	return row


def read_server_csv_files(server_names):
	base_dir = Path(__file__).resolve().parent
	input_files = list(base_dir.glob("*.csv")) + list(base_dir.glob("*.txt"))
	data_by_server = {}
	missing_servers = []

	for server in server_names:
		# Prefer exact filename matches: <SERVER>.csv, then <SERVER>.txt
		matched_files = []
		for suffix in (".csv", ".txt"):
			exact_path = base_dir / f"{server}{suffix}"
			if exact_path.exists():
				matched_files = [exact_path]
				break

		# Fallback: any supported input filename containing the server name.
		if not matched_files:
			matched_files = [p for p in input_files if server.lower() in p.stem.lower()]

		if not matched_files:
			missing_servers.append(server)
			continue

		rows = []
		fieldnames = None
		yesterday = (datetime.now() - timedelta(days=1)).date()
		for path in matched_files:
			with path.open("r", newline="", encoding="utf-8-sig") as f:
				reader = csv.DictReader(f)
				if fieldnames is None and reader.fieldnames:
					fieldnames = list(reader.fieldnames)
				for row in reader:
					row = _merge_task_category_overflow(row)
					row_date = _get_row_date(row)
					event_id = _get_event_id(row)
					# Keep only rows where Date and Time is yesterday and Event ID is not in the whitelist.
					if row_date == yesterday and event_id not in Error_whitelist:
						rows.append(row)

		data_by_server[server] = {
			"files": [str(p.name) for p in matched_files],
			"fieldnames": fieldnames or [],
			"rows": rows,
		}

	return data_by_server, missing_servers


def write_filtered_csv_files(server_data):
	base_dir = Path(__file__).resolve().parent
	output_dir = base_dir / "csv_filtered"
	output_dir.mkdir(parents=True, exist_ok=True)
	written_files = {}

	for server, info in server_data.items():
		output_path = output_dir / f"{server}_filtered.csv"
		fieldnames = list(info.get("fieldnames") or [])
		rows = info.get("rows", [])

		if not fieldnames and rows:
			fieldnames = list(rows[0].keys())

		with output_path.open("w", newline="", encoding="utf-8-sig") as f:
			if fieldnames:
				clean_fieldnames = [name for name in fieldnames if name is not None]
				writer = csv.DictWriter(f, fieldnames=clean_fieldnames, extrasaction="ignore")
				writer.writeheader()
				if rows:
					writer.writerows(rows)

		written_files[server] = str(output_path.relative_to(base_dir))

	return written_files


def read_filtered_csv_summaries():
	base_dir = Path(__file__).resolve().parent
	output_dir = base_dir / "csv_filtered"
	summaries = {}

	if not output_dir.exists():
		return summaries

	for path in output_dir.glob("*_filtered.csv"):
		server_name = path.stem.removesuffix("_filtered")
		row_count = 0
		event_ids = set()
		with path.open("r", newline="", encoding="utf-8-sig") as f:
			reader = csv.DictReader(f)
			for row in reader:
				row_count += 1
				event_id = _get_event_id(row)
				if event_id is not None:
					event_ids.add(event_id)

		summaries[server_name] = {
			"row_count": row_count,
			"event_ids": sorted(event_ids),
		}

	return summaries


def _get_task_category_value(row):
	for key, value in row.items():
		if isinstance(key, str) and key.strip().lower() == "task category":
			task_value = str(value or "").strip()
			return "" if task_value.lower() == "none" else task_value
	return ""


def build_log_message_rows_from_filtered_csv():
	base_dir = Path(__file__).resolve().parent
	output_dir = base_dir / "csv_filtered"
	grouped_rows = []

	if not output_dir.exists():
		return grouped_rows

	for path in get_filtered_csv_paths_in_template_order(output_dir):
		server_name = path.stem.removesuffix("_filtered")
		ordered_groups = {}
		seen_1377_scope_suffixes = set()

		with path.open("r", newline="", encoding="utf-8-sig") as f:
			reader = csv.DictReader(f)
			for row in reader:
				event_id = _get_event_id(row)
				if event_id is None:
					continue

				task_value = _get_task_category_value(row)
				if event_id == 1377:
					scope_suffix = task_value
					lower_task_value = task_value.lower()
					marker = "scopes"
					if marker in lower_task_value:
						marker_index = lower_task_value.index(marker)
						scope_suffix = task_value[marker_index + len(marker):].strip()

					normalized_scope_suffix = scope_suffix.lower()
					if normalized_scope_suffix in seen_1377_scope_suffixes:
						continue
					seen_1377_scope_suffixes.add(normalized_scope_suffix)

				group_key = (event_id, task_value)
				if group_key not in ordered_groups:
					ordered_groups[group_key] = 0
				ordered_groups[group_key] += 1

		if not ordered_groups:
			continue

		server_rows = []
		for (event_id, task_value), count in ordered_groups.items():
			display_task = task_value or "(blank)"
			if count > 1:
				display_task = f"({count} times) {display_task}"
			server_rows.append({
				"server_name": server_name,
				"event_id": event_id,
				"task_category": display_task,
			})

		grouped_rows.append(server_rows)

	return grouped_rows


def update_log_message_xlsx_from_filtered_csv():
	base_dir = Path(__file__).resolve().parent
	log_message_path = base_dir / "Log_message.xlsx"

	if not log_message_path.exists():
		print(f"Log workbook not found: {log_message_path.name}")
		return False

	grouped_rows = build_log_message_rows_from_filtered_csv()

	wb = load_workbook(log_message_path)
	ws = wb.active

	for merged_range in list(ws.merged_cells.ranges):
		if merged_range.min_col <= 3 and merged_range.max_col >= 1:
			ws.unmerge_cells(str(merged_range))

	# Remove inherited border styles so no thick outside border remains.
	for row in range(1, ws.max_row + 1):
		for col in range(1, ws.max_column + 1):
			cell = ws.cell(row=row, column=col)
			if isinstance(cell, MergedCell):
				continue
			cell.border = Border()

	# Clear existing output values in columns A-C while keeping workbook formatting.
	for row in range(1, ws.max_row + 1):
		for col in range(1, 4):
			cell = ws.cell(row=row, column=col)
			if isinstance(cell, MergedCell):
				continue
			cell.value = None

	write_row = 1
	for server_rows in grouped_rows:
		first_for_server = True
		previous_event_id = None

		for item in server_rows:
			server_name = item["server_name"]
			event_id = item["event_id"]
			task_category = item["task_category"]

			ws.cell(row=write_row, column=1).value = server_name if first_for_server else None
			ws.cell(row=write_row, column=2).value = event_id if (first_for_server or event_id != previous_event_id) else None
			ws.cell(row=write_row, column=3).value = task_category

			first_for_server = False
			previous_event_id = event_id
			write_row += 1

		# Keep one blank row between servers to match the reference layout style.
		write_row += 1

	thin_side = Side(style="thin", color="000000")
	full_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
	calibri_font = Font(name="Calibri")
	center_alignment = Alignment(horizontal="center", vertical="center")

	for row in range(1, ws.max_row + 1):
		for col in range(1, ws.max_column + 1):
			cell = ws.cell(row=row, column=col)
			if isinstance(cell, MergedCell):
				continue
			cell.font = calibri_font
			if col in (1, 2):
				cell.alignment = center_alignment

	for row in range(1, ws.max_row + 1):
		for col in range(1, ws.max_column + 1):
			cell = ws.cell(row=row, column=col)
			if isinstance(cell, MergedCell):
				continue
			value = cell.value
			if value is None or str(value).strip() == "":
				continue
			cell.border = full_border

	wb.save(log_message_path)
	print(f"Updated {log_message_path.name} from csv_filtered files.")
	return True


def get_template_server_sequence():
	base_dir = Path(__file__).resolve().parent
	template_path = base_dir / "DHCP_daily_template.xlsx"

	if not template_path.exists():
		return []

	wb = load_workbook(template_path, read_only=True)
	ws = wb.active
	server_names = []
	for row in range(3, 25):
		server_name = str(ws.cell(row=row, column=3).value or "").strip()
		if server_name:
			server_names.append(server_name)
	return server_names


def get_filtered_csv_paths_in_template_order(output_dir):
	path_by_server = {
		path.stem.removesuffix("_filtered"): path
		for path in output_dir.glob("*_filtered.csv")
	}
	ordered_paths = []
	seen_servers = set()

	for server_name in get_template_server_sequence():
		path = path_by_server.get(server_name)
		if path is None:
			continue
		ordered_paths.append(path)
		seen_servers.add(server_name)

	for server_name, path in sorted(path_by_server.items()):
		if server_name in seen_servers:
			continue
		ordered_paths.append(path)

	return ordered_paths


def print_task_category_counts_for_1063():
	base_dir = Path(__file__).resolve().parent
	output_dir = base_dir / "csv_filtered"

	if not output_dir.exists():
		print("csv_filtered folder not found.")
		return

	print("Task Category counts for files containing Event ID 1063:")
	for path in get_filtered_csv_paths_in_template_order(output_dir):
		task_counts = {}

		with path.open("r", newline="", encoding="utf-8-sig") as f:
			reader = csv.DictReader(f)
			for row in reader:
				event_id = _get_event_id(row)
				if event_id != 1063:
					continue

				task_value = ""
				for key, value in row.items():
					if isinstance(key, str) and key.strip().lower() == "task category":
						task_value = str(value or "").strip()
						break
				if task_value.lower() == "none":
					task_value = ""
				if not task_value:
					task_value = "(blank)"
				task_counts[task_value] = task_counts.get(task_value, 0) + 1

		if not task_counts:
			continue

		print(f"- {path.name}")
		for task_value, count in sorted(task_counts.items(), key=lambda item: (-item[1], item[0])):
			print(f"    {count} | {task_value}")


def print_rows_for_1377():
	base_dir = Path(__file__).resolve().parent
	output_dir = base_dir / "csv_filtered"

	if not output_dir.exists():
		print("csv_filtered folder not found.")
		return

	print("Rows with Event ID 1377 in csv_filtered files:")
	printed_section = False
	for path in get_filtered_csv_paths_in_template_order(output_dir):
		rows_1377 = []
		seen_scope_suffixes = set()

		with path.open("r", newline="", encoding="utf-8-sig") as f:
			reader = csv.DictReader(f)
			for row in reader:
				if _get_event_id(row) != 1377:
					continue

				task_value = ""
				for key, value in row.items():
					if isinstance(key, str) and key.strip().lower() == "task category":
						task_value = str(value or "").strip()
						break

				scope_suffix = task_value
				lower_task_value = task_value.lower()
				marker = "scopes"
				if marker in lower_task_value:
					marker_index = lower_task_value.index(marker)
					scope_suffix = task_value[marker_index + len(marker):].strip()

				normalized_scope_suffix = scope_suffix.lower()
				if normalized_scope_suffix in seen_scope_suffixes:
					continue

				seen_scope_suffixes.add(normalized_scope_suffix)
				rows_1377.append(row)

		if not rows_1377:
			continue

		if printed_section:
			print()
			print()

		print(f"- {path.name}")
		for idx, row in enumerate(rows_1377, start=1):
			print(f"    Row {idx}: {row}")
		printed_section = True


def update_template_values(filtered_summaries):
	base_dir = Path(__file__).resolve().parent
	template_path = base_dir / "DHCP_daily_template.xlsx"

	if not template_path.exists():
		print(f"Template file not found: {template_path.name}")
		return False

	wb = load_workbook(template_path)
	ws = wb.active
	no_fill = PatternFill(fill_type=None)
	yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")

	for row in range(2, 25):
		for col in range(1, 8):
			ws.cell(row=row, column=col).fill = no_fill

	for row in range(3, 25):
		ws.cell(row=row, column=4).value = "Normal"
		ws.cell(row=row, column=7).value = None

	for row in range(3, 25):
		server_name = str(ws.cell(row=row, column=3).value or "").strip()
		summary = filtered_summaries.get(server_name)
		if not summary or summary["row_count"] <= 0:
			continue

		event_id_text = ", ".join(str(event_id) for event_id in summary["event_ids"])
		ws.cell(row=row, column=4).value = "Abnormal"
		ws.cell(row=row, column=7).value = f"Error: {event_id_text}" if event_id_text else "Error:"
		ws.cell(row=row, column=4).fill = yellow_fill
		ws.cell(row=row, column=7).fill = yellow_fill

	wb.save(template_path)
	print(f"Updated {template_path.name}: reset A2:G24 fill, set D3:D24 to 'Normal', cleared G3:G24, and marked rows with filtered errors")
	return True


if __name__ == "__main__":
	server_data, missing = read_server_csv_files(Have_Log_Server_list)
	written = write_filtered_csv_files(server_data)
	filtered_summaries = read_filtered_csv_summaries()
	update_template_values(filtered_summaries)
	update_log_message_xlsx_from_filtered_csv()
	print_task_category_counts_for_1063()
	print_rows_for_1377()

	print(f"Servers in list: {len(Have_Log_Server_list)}")
	print(f"Servers with CSV data: {len(server_data)}")
    
	for server, info in server_data.items():
		print(f"{server}: {len(info['rows'])} rows from {', '.join(info['files'])}")
		print(f"  -> wrote {written[server]}")

	if missing:
		print("Missing CSV for:", ", ".join(missing))
  
